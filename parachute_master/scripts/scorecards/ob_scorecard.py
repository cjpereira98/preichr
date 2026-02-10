import datetime
import shutil
import openpyxl
import os
import sys

# Add the parent directory to the system path to allow importing integrations within ampy
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../../src')))
from ampy.metrics.ob_diverts import OBDiverts
from ampy.metrics.rc_sort_total_rate import RCSortTotalRate
from ampy.metrics.lane_full import LaneFull
from ampy.metrics.fl import FL
from ampy.metrics.mp import MP
from ampy.metrics.rwc import RWC
from ampy.metrics.manual_edit_hours import ManualEditHours
from ampy.metrics.pallets_loaded import PalletsLoaded
from ampy.integrations.firefox import FirefoxIntegration

# List of metric class names (as strings)
metric_names = [
    'OBDiverts', 'RCSortTotalRate', 'LaneFull', 'FL', 
    'MP', 'RWC', 'ManualEditHours', 'PalletsLoaded'
]

# Copy the template file to a new file for the previous business week and set week number in D2
def create_weekly_scorecard():
    # Get the previous business week number
    today = datetime.date.today()
    last_sunday = today - datetime.timedelta(days=today.weekday() + 1)
    week_num = last_sunday.isocalendar()[1]
    
    template_path = 'OB_SCORECARD_TEMPLATE.xlsx'
    new_file_path = f'OB_SCORECARD_WEEK{week_num}.xlsx'
    
    # Copy the template file
    shutil.copy(template_path, new_file_path)
    
    # Open the new workbook and set "Week XX" in D2
    wb = openpyxl.load_workbook(new_file_path)
    ws = wb.active
    ws['D2'] = f"Week {week_num}"
    
    # Save and close the workbook
    wb.save(new_file_path)
    
    return new_file_path

# Populate metrics by dynamically creating instances and using their get method
def populate_metrics(ws, driver):
    # Set ranges for metric rows and columns
    shifts = {'FHD': 5, 'BHD': 7, 'FHN': 9, 'BHN': 11}
    metric_columns = list('DEFGHIJKL')
    
    # Dynamically create instances of each metric class with the shared driver
    metric_instances = {}
    for metric_name in metric_names:
        metric_class = globals().get(metric_name)
        if metric_class:
            metric_instances[metric_name] = metric_class(driver=driver)

    # Populate metrics on the worksheet for each shift
    for col_letter, metric_name in zip(metric_columns, metric_names):
        metric_instance = metric_instances.get(metric_name)
        if metric_instance:
            for shift, row in shifts.items():
                cell = f"{col_letter}{row}"
                # Use shift-specific specificity for each metric
                ws[cell] = metric_instance.get(shift)  

# Function to calculate rankings based on metrics
def calculate_rankings(ws):
    # Columns with metrics (D-K)
    metric_columns = list('DEFGHIJK')
    shifts = {'FHD': 5, 'BHD': 7, 'FHN': 9, 'BHN': 11}
    rank_rows = {'FHD': 6, 'BHD': 8, 'FHN': 10, 'BHN': 12}

    # Calculate individual metric rankings by column
    for col_letter in metric_columns:
        # Get shift values for this column
        values = [(shift, ws[f"{col_letter}{shifts[shift]}"].value) for shift in shifts]
        
        # Sort by metric: J (lowest wins), all others (highest wins)
        if col_letter == 'J':
            sorted_values = sorted(values, key=lambda x: x[1])
        else:
            sorted_values = sorted(values, key=lambda x: x[1], reverse=True)
        
        # Assign ranks (1-4)
        for rank, (shift, value) in enumerate(sorted_values, start=1):
            rank_cell = f"{col_letter}{rank_rows[shift]}"
            ws[rank_cell] = rank

    # Sum the ranks for each shift and place in column L, calculate final rank
    total_ranks = []
    for shift, row in shifts.items():
        # Calculate the sum of ranks for this shift
        total_rank = sum([ws[f"{col}{row+1}"].value for col in metric_columns])
        ws[f"L{row}"] = total_rank  # Place total in column L
        
        # Store the total rank for overall ranking
        total_ranks.append((shift, total_rank))

    # Determine the final ranking (1-4) based on total score and place in column M
    sorted_totals = sorted(total_ranks, key=lambda x: x[1])
    for rank, (shift, total) in enumerate(sorted_totals, start=1):
        ws[f"M{shifts[shift]}"] = rank

# Main function to execute the script
def main():
    # Authenticate and create a Firefox driver
    driver = FirefoxIntegration.get_authenticated_driver()

    # Create a new scorecard for the previous business week
    new_file_path = create_weekly_scorecard()
    
    # Open the workbook and select the active sheet
    wb = openpyxl.load_workbook(new_file_path)
    ws = wb.active
    
    # Populate metrics using reflection logic with the authenticated driver
    populate_metrics(ws, driver)
    
    # Calculate rankings based on populated metrics
    calculate_rankings(ws)
    
    # Save the workbook
    wb.save(new_file_path)
    print(f"Scorecard created and saved as {new_file_path}")

    # Close the driver after the script completes
    driver.quit()

# Run the script
if __name__ == "__main__":
    main()
