# src/ampy/integrations/apollo.py
import sys
import os
import time
import csv
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.firefox.options import Options

# Add the parent directory to the system path to allow importing FirefoxIntegration
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

# Add the directory three levels up to the system path to allow importing FirefoxIntegration
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', '..')))

from config.config import OWNER

# Add the location of openpyxl to the system path
openpyxl_path = f"C:\\Users\\{OWNER}\\AppData\\Roaming\\Python\\Python310\\site-packages\\openpyxl"
sys.path.append(openpyxl_path)


from openpyxl import load_workbook
from integrations.firefox import FirefoxIntegration

class ApolloIntegration:
    @staticmethod
    def get_audit_results(audit_id, start_date, end_date):
        """
        Retrieves audit results for a specific date range and converts them into a CSV format.
        
        :param audit_id: The ID of the audit to retrieve results for.
        :param start_date: The start date for the audit results (in 'YYYY-MM-DD HH:MM:SS +/-HHMM' format).
        :param end_date: The end date for the audit results (in 'YYYY-MM-DD HH:MM:SS +/-HHMM' format).
        :return: A list of dictionaries containing the audit results.
        """

        # Set up Firefox profile to handle downloads
        download_dir = os.getcwd()
        options = Options()
        profile = webdriver.FirefoxProfile()
        profile.set_preference("browser.download.folderList", 2)
        profile.set_preference("browser.download.manager.showWhenStarting", False)
        profile.set_preference("browser.download.dir", download_dir)
        profile.set_preference("browser.helperApps.neverAsk.saveToDisk", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


        driver = FirefoxIntegration.get_authenticated_driver()

        # Format the URL with the given parameters
        url = f"https://apollo-audit.corp.amazon.com/reporting/results_by_audit?audit_type_id={audit_id}&end_date={end_date}&start_date={start_date}"
        driver.get(url)
        
        # Wait for the download button to be visible and click it
        try:
            download_button = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, "//a[contains(@href, 'results_by_audit.xlsx') and contains(@class, 'btn-primary')]"))
            )
            download_button.click()
        except Exception as e:
            print(f"Error: {e}")
            driver.quit()
            return []
        
    

        # Wait for the download to complete (adjust time as necessary)
        time.sleep(5)

        # Locate the downloaded file (assuming default Firefox download directory)
        download_dir = os.getcwd()
        file_path = os.path.join(download_dir, "results_by_audit.xlsx")

        # Convert Excel to CSV
        csv_file_path = os.path.join(download_dir, "results_by_audit.csv")
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active

        with open(csv_file_path, mode='w', newline="", encoding='utf-8') as csv_file:
            csv_writer = csv.writer(csv_file)
            for row in sheet.iter_rows(values_only=True):
                csv_writer.writerow(row)

        # Clean up the downloaded file
        os.remove(file_path)
        driver.quit()

        # Process the CSV file
        results = []
        with open(csv_file_path, mode='r', encoding='utf-8') as csv_file:
            csv_reader = csv.DictReader(csv_file)
            for row in csv_reader:
                results.append(row)

        # Clean up the CSV file
        os.remove(csv_file_path)

        return results