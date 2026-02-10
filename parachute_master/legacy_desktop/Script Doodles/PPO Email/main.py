import json
import requests
import csv
import time
import os
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pandas as pd
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo
from seleniumwire import webdriver as selenium_wire_webdriver

# CONTROLS
# test mode
test_mode = False  # Set to True for testing, will slow down execution

# slackbot url
#SLACKBOT_URL = "https://hooks.slack.com/triggers/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"

# speed
NEO_QUERY_SPEED = 2 #seconds waiting for each Neo query to complete (roughly 152 queries in present state)
NEO_QUERY_SPEED_2 = 30

# sites
#SITES = {"GYR3", "IND9", "LAS1", "LAX9", "MEM1", "MQJ1", "SMF3", "TEB9", "FWA4",
#         "GYR2", "IAH3", "ORF2", "RDU2", "RFD2", "RMN3", "SBD1", "SCK4", "SWF2", "VGT2", "AVP1", "ABE8", "CLT2", "FTW1", "MDW2", "LGB8", "ONT8"}
#SITES = {"LAX9"}


# helpers

def create_metrics_dict():
    """
    Creates a dictionary of sites with their respective staffing data.
    """
    sites_dict = {}
    sites_dict = {
        "fhd_trailer_audits": 0,
        "bhd_trailer_audits": 0,
        "fhn_trailer_audits": 0,
        "bhn_trailer_audits": 0,
        "total_trailer_audits": 0,
        "dwells_over_eight": 0,
        "lp_scan_compliance": 0,
        "fhd_pid_cartons": 0,
        "fhn_pid_cartons": 0,
        "bhd_pid_cartons": 0,
        "bhn_pid_cartons": 0,
        "total_pid_cartons": 0,
        "total_combined_cartons": 0,
        "total_pr_pallets": 0,
        "total_ti_volume": 0,
        "total_offline_cartons": 0,
        "ib_pr_cpp": 0,
        "rd_cplh_incl_ti": 0,
        "total_ib_cplh": 0,
        }
    return sites_dict






def pull_actuals(sites_dict, driver):
    processes = {
        "Receive Dock": 1003010,
        "Pallet-Receive": 1003032,
        "LP-Receive": 1003031,
        "Transfer In Support": 1003020
    }

    tz_df = pd.read_csv("timezones.csv")
    tz_lookup = dict(zip(tz_df["Site"], tz_df["Delta to EST"]))

    yesterday_6am = (datetime.now() - timedelta(days=1)).replace(hour=6, minute=0, second=0, microsecond=0)

    for site in sites_dict:
        site_delta = tz_lookup.get(site, 0)
        local_yesterday_6am = yesterday_6am
        start_date = local_yesterday_6am
        end_date = start_date + timedelta(days=1)
        start_hour = start_date.hour
        start_minute = start_date.minute
        end_hour = end_date.hour
        end_minute = end_date.minute

        start_date_str = start_date.strftime("%Y/%m/%d")
        end_date_str = end_date.strftime("%Y/%m/%d")
        start_dt_local = datetime.combine(start_date.date(), datetime.min.time()) + timedelta(hours=start_hour, minutes=start_minute)
        end_dt_local = datetime.combine(end_date.date(), datetime.min.time()) + timedelta(hours=end_hour, minutes=end_minute)
        start_dt_utc = (start_dt_local - timedelta(hours=site_delta)).replace(tzinfo=ZoneInfo("America/New_York")).astimezone(ZoneInfo("UTC")).strftime("%Y%m%d%H%M%S")
        end_dt_utc = (end_dt_local - timedelta(hours=site_delta)).replace(tzinfo=ZoneInfo("America/New_York")).astimezone(ZoneInfo("UTC")).strftime("%Y%m%d%H%M%S")

        total_hours = 0
        excluded_hours = 0
        legacy_only_hours = 0

        for process_name, process_id in processes.items():
            success = False
            for attempt in range(3):
                url = (
                    f"https://fclm-portal.amazon.com/reports/functionRollup?"
                    f"reportFormat=CSV&warehouseId={site}&processId={process_id}&maxIntradayDays=1"
                    f"&spanType=Intraday&startDateIntraday={start_date_str}&startHourIntraday={start_hour}"
                    f"&startMinuteIntraday={start_minute}&endDateIntraday={end_date_str}&endHourIntraday={end_hour}"
                    f"&endMinuteIntraday={end_minute}"
                )

                driver.execute_script(f"window.open('{url}', '_blank');")

                filename = f"functionRollupReport-{site}-{process_name}-Intraday-{start_dt_utc}-{end_dt_utc}.csv"
                filepath = os.path.join(os.getcwd(), filename)

                if test_mode:
                    print(f"[INFO] Attempt {attempt + 1}: Downloading {filename} from {url}")

                for _ in range(15):
                    if os.path.exists(filepath):
                        print(f"[INFO] Downloaded {filename}")
                        if test_mode:
                            time.sleep(5)
                        break
                    time.sleep(1)

                if not os.path.exists(filepath):
                    print(f"[WARN] Attempt {attempt + 1}: File did not download: {filename}")
                    continue

                try:
                    time.sleep(3)  # Allow time for file to be fully written
                    df = pd.read_csv(filepath)
                    if df.empty or "Employee Id" not in df.columns or "Function Name" not in df.columns:
                        print(f"[WARN] Missing required data in {filename}")
                        continue

                    dedup_df = df.drop_duplicates(subset=["Employee Id", "Function Name"])
                    
                    if "Paid Hours-Total(function,employee)" in dedup_df.columns:
                        # Identify excluded hours before filtering
                        excluded_df = dedup_df[dedup_df["Function Name"].isin([
                            "Bin Scout Trans", "Cart Runner Trans", "Tote Stacking", 
                            "Transfer In Ambssdr", "Transfer In Training", "PrEditor Receive"
                        ])]
                        excluded_hours += excluded_df["Paid Hours-Total(function,employee)"].sum()

                        legacy_only_df = dedup_df[dedup_df["Function Name"].isin([
                            "PID Manual Divert", "Vendor Cutter"
                        ])]
                        legacy_only_hours += legacy_only_df["Paid Hours-Total(function,employee)"].sum()

                        # Now filter them out
                        dedup_df = dedup_df[~dedup_df["Function Name"].isin([
                            "Bin Scout Trans", "Cart Runner Trans", "Tote Stacking", 
                            "Transfer In Ambssdr", "Transfer In Training", "PrEditor Receive", "PID Manual Divert", "Vendor Cutter"
                        ])]

                        hours = dedup_df["Paid Hours-Total(function,employee)"].sum()
                        total_hours += hours
                        success = True
                        print(f"[INFO] Processed {filename}")
                    else:
                        print(f"[WARN] 'Paid Hours' column not found in {filename}")

                except Exception as e:
                    print(f"[ERROR] Could not process {filename}: {e}")

                finally:
                    if os.path.exists(filepath):
                        print(f"[INFO] Deleting {filename}")
                        os.remove(filepath)

                if success:
                    break

            if not success:
                print(f"[FAIL] All attempts failed for {site} - {process_name}")

        sites_dict[site]["actual staffing"] = round(total_hours)
        sites_dict[site]["excluded_hours"] = round(excluded_hours)
        sites_dict[site]["legacy_only_hours"] = round(legacy_only_hours)
        print(f"[INFO] {site} - Actual Staffing: {sites_dict[site]['actual staffing']} hours, Excluded: {sites_dict[site]['excluded_hours']} hours")

    return sites_dict

def pull_trailer_audits(sites_dict, driver):
    today = datetime.today()
    most_recent_sunday = today - timedelta(days=today.weekday() + 1 if today.weekday() != 6 else 0)
    start_date = (most_recent_sunday - timedelta(days=7)).strftime("%Y-%m-%d")
    end_date = most_recent_sunday.strftime("%Y-%m-%d")

    audit_id = 23770

    # Format the URL with the given parameters
    url = f"https://apollo-audit.corp.amazon.com/reporting/results_by_audit?audit_type_id={audit_id}&end_date={end_date}&start_date={start_date}"
    driver.get(url)
    
    # Wait for the download button to be visible and click it
    try:
        download_button = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, "//a[contains(@href, 'results_by_audit.xlsx') and contains(@class, 'btn-primary')]"))
        )
        download_button.click()
    except Exception as e:
        print(f"Error: {e}")
        return sites_dict
    


    # Wait for the download to complete (adjust time as necessary)
    time.sleep(5)

    # Locate the downloaded file (assuming default Firefox download directory)
    download_dir = os.getcwd()
    target_prefix = "results_by_audit"
    target_ext = ".xlsx"

    def find_latest_xlsx():
        cands = [
            os.path.join(download_dir, f)
            for f in os.listdir(download_dir)
            if f.startswith(target_prefix) and f.endswith(target_ext)
        ]
        if not cands:
            return None
        return max(cands, key=lambda p: os.path.getmtime(p))

    # Poll up to 60s for a new/updated file
    deadline = time.time() + 60
    xlsx_path = None
    while time.time() < deadline:
        candidate = find_latest_xlsx()
        if candidate and os.path.isfile(candidate):
            # Basic "download finished" heuristic: file not growing across two checks
            size1 = os.path.getsize(candidate)
            time.sleep(0.5)
            size2 = os.path.getsize(candidate)
            if size1 == size2 and size1 > 0:
                xlsx_path = candidate
                break
        time.sleep(0.5)

    if not xlsx_path:
        print("[pull_trailer_audits] Downloaded file not found or incomplete.")
        return sites_dict

    # --- Parse workbook directly to avoid CSV conversion loss ---
    try:
        wb = load_workbook(filename=xlsx_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"[pull_trailer_audits] Failed to open workbook: {e}")
        return sites_dict

    # Aggregate counters
    counts = {"FHD": 0, "BHD": 0, "FHN": 0, "BHN": 0}

    def normalize(s):
        return (s or "").strip()

    # Iterate all sheets that start with "results" (ignore "PLEASE READ" and others)
    for ws in wb.worksheets:
        if not ws.title.lower().startswith("results"):
            continue

        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            continue  # empty sheet

        if not header:
            continue

        # Build header index by name; we’ll prefer name lookup, but fall back to Excel letters if needed
        name_to_idx = {normalize(str(h)): idx for idx, h in enumerate(header)}

        # Expected headers per prompt
        role_col_name = "Auditor Role"
        shift_col_name = "Current Shift"

        role_idx = name_to_idx.get(role_col_name, None)
        shift_idx = name_to_idx.get(shift_col_name, None)

        # If either column missing, skip this sheet gracefully
        if role_idx is None or shift_idx is None:
            # Optional: try fallback by absolute columns if you’re 100% certain:
            # role_idx = 4  # column E (0-based)
            # shift_idx = 6 # column G (0-based)
            # If still None, continue
            continue

        for row in rows_iter:
            if not row:
                continue

            role_val = normalize(row[role_idx]).upper()
            shift_val = normalize(row[shift_idx]).upper()

            # Filter to OM/AM only
            if role_val == "OM/AM":
                if shift_val in counts:
                    counts[shift_val] += 1

    total = sum(counts.values())

    time.sleep(5)
    # Optional: clean up downloaded file
    print(xlsx_path)
    #try:
    wb.close()
    os.remove(xlsx_path)
    #except Exception:
    #    pass

    # Push counts into every site entry (global totals). If you need per-site,
    # map by a Site/FC column from the report and aggregate by key instead.
    
    sites_dict["fhd_trailer_audits"] = counts["FHD"]
    sites_dict["bhd_trailer_audits"] = counts["BHD"]
    sites_dict["fhn_trailer_audits"] = counts["FHN"]
    sites_dict["bhn_trailer_audits"] = counts["BHN"]
    sites_dict["total_trailer_audits"] = total

    return sites_dict

def pull_carton_data(sites_dict, driver):
    return sites_dict

def output_to_csv(sites_dict):
    """Outputs the keys and values from sites_dict into a two-column CSV file."""
    output_file = "ppo_output.csv"
    with open(output_file, mode="w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerow(["Key", "Value"])
        for key, value in sites_dict.items():
            writer.writerow([key, value])
    print(f"✅ Output written to {output_file}")

# main

def main():

    # Launch browser for manual authentication
    print("Launching Firefox for Midway auth... you have 20 seconds.")
    options = Options()
    options.set_preference("browser.download.folderList", 2)
    options.set_preference("browser.download.dir", os.getcwd())
    options.set_preference("browser.helperApps.neverAsk.saveToDisk", "text/csv")
    #options.add_argument("--auto-open-devtools-for-tabs")
    driver = selenium_wire_webdriver.Firefox(options=options)
    #driver.execute_cdp_cmd("Network.enable",{})
    driver.get("https://atoz.amazon.work/")  # Or the exact URL you want to pre-auth

    #time.sleep(7200)
    time.sleep(30)  # Wait for manual login and cert push
    print("Auth window complete. Continuing...")

    sites_dict = create_metrics_dict()
    sites_dict = pull_trailer_audits(sites_dict, driver)
    sites_dict = pull_carton_data(sites_dict, driver)

    #sites_dict = calculate_recommended_staffing(sites_dict, driver)
    #sites_dict = pull_actuals(sites_dict, driver)
    #sites_dict = calculate_deltas(sites_dict)
    
    output_to_csv(sites_dict)



if __name__ == "__main__":
    main()